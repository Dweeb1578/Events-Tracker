from agents import excel_writer
from agents import sheets_writer


def test_excel_engager_columns_include_new_fields():
    for col in ("Verdict", "Engagement Strength", "Event Date", "Event City", "Warm Reason"):
        assert col in excel_writer.ENGAGER_COLUMNS


def test_sheets_engager_headers_include_new_fields():
    for col in ("Verdict", "Engagement Strength", "Event Date", "Event City", "Warm Reason"):
        assert col in sheets_writer.ENGAGER_HEADERS


def test_excel_writes_engager_rows(tmp_path):
    out = tmp_path / "e.xlsx"
    n = excel_writer.write_engagers([{
        "name": "Jane Doe", "linkedin_url": "https://linkedin.com/in/ACoAA1",
        "title": "VP Finance", "company": "AcmeSaaS", "icp_score": 86,
        "verdict": "STRONG ICP FIT", "engagement_type": "commenter",
        "engagement_strength": "high (comment)", "event_date": "2026-07-15",
        "event_city": "NYC", "warm_reason": "Commented on Orb's post (NYC)",
        "source_post_url": "p1", "source_post_company": "Orb",
    }], output_path=str(out))
    assert n == 1 and out.exists()


def test_excel_engager_escapes_formula_injection(tmp_path):
    from openpyxl import load_workbook
    out = tmp_path / "evil.xlsx"
    excel_writer.write_engagers([{
        "name": '=HYPERLINK("http://evil")', "linkedin_url": "https://linkedin.com/in/x",
        "title": "VP Finance", "company": "AcmeSaaS", "icp_score": 86,
    }], output_path=str(out))
    wb = load_workbook(str(out))
    ws = wb["Engagers"]
    name_col = excel_writer.ENGAGER_COLUMNS.index("Name") + 1
    assert ws.cell(row=2, column=name_col).value == '\'=HYPERLINK("http://evil")'
