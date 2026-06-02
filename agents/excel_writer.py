"""
Agent 4: Deduplication + Excel Writer Agent
Writes data to Excel with two sheets:
  - "Scraped Data": raw scraper output
  - "Classified Events": LLM-classified events with formatting
"""

import os
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "output")
DEFAULT_OUTPUT = os.path.join(OUTPUT_DIR, "events.xlsx")

# ── Classified Events sheet columns ──
EVENT_COLUMNS = [
    "Event Name",
    "Host Company",
    "Category",
    "Event Type",
    "Date",
    "Location",
    "Target Audience",
    "Registration URL",
    "Description",
    "Relevance Score",
    "Source URL",
    "Detected At",
    "Status",
]

# ── Scraped Data sheet columns ──
SCRAPED_COLUMNS = [
    "Company",
    "Category",
    "Source URL",
    "Source Type",
    "Scrape Method",
    "Scraped At",
    "Content Length",
    "Content Preview",
]

# Color codes for event types
EVENT_TYPE_COLORS = {
    "summit": "4472C4",
    "conference": "4472C4",
    "dinner": "E2725B",
    "roundtable": "ED7D31",
    "webinar": "70AD47",
    "workshop": "FFC000",
    "meetup": "9DC3E6",
    "other": "A5A5A5",
}

# Styles
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SCRAPED_HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _format_sheet(ws, columns, header_fill):
    """Apply header formatting to a worksheet."""
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"

    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"


def _set_column_widths(ws, widths: dict, columns: list):
    """Set column widths for a worksheet."""
    for col_idx, col_name in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = widths.get(col_name, 15)


def _get_existing_keys(ws, name_col: int, company_col: int, date_col: int) -> set[tuple]:
    """Get set of dedup keys from existing sheet."""
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[name_col]:
            key = (
                str(row[name_col]).strip().lower(),
                str(row[company_col]).strip().lower() if row[company_col] else "",
                str(row[date_col]).strip().lower() if row[date_col] else "",
            )
            existing.add(key)
    return existing


def _color_event_type(ws, row_idx: int, event_type: str, columns: list):
    """Apply color coding to event type cell."""
    event_type_lower = event_type.lower() if event_type else "other"
    color = EVENT_TYPE_COLORS.get(event_type_lower, EVENT_TYPE_COLORS["other"])
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font_color = "FFFFFF" if color in ("1F4E79", "4472C4", "E2725B") else "000000"

    type_col = columns.index("Event Type") + 1
    cell = ws.cell(row=row_idx, column=type_col)
    cell.fill = fill
    cell.font = Font(name="Calibri", color=font_color, bold=True, size=10)


# ─────────────────────────────────────────────────────────
# Public API: write scraped data and classified events
# ─────────────────────────────────────────────────────────

def write_scraped_data(scraped_items: list[dict], output_path: str = DEFAULT_OUTPUT) -> int:
    """
    Write raw scraped data to the 'Scraped Data' sheet.
    Returns number of rows written.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    if os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        # Remove default sheet if creating fresh
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Get or create scraped data sheet
    sheet_name = "Scraped Data"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        _format_sheet(ws, SCRAPED_COLUMNS, SCRAPED_HEADER_FILL)
        _set_column_widths(ws, {
            "Company": 18, "Category": 15, "Source URL": 45,
            "Source Type": 18, "Scrape Method": 14, "Scraped At": 22,
            "Content Length": 14, "Content Preview": 60,
        }, SCRAPED_COLUMNS)

    count = 0
    for item in scraped_items:
        row_idx = ws.max_row + 1
        raw_text = item.get("raw_text", "")
        row_data = [
            item.get("company", ""),
            item.get("category", ""),
            item.get("source_url", ""),
            item.get("source_type", ""),
            item.get("scrape_method", ""),
            item.get("scraped_at", ""),
            len(raw_text),
            raw_text[:200] + "..." if len(raw_text) > 200 else raw_text,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == len(row_data)))
            cell.font = Font(name="Calibri", size=10)

        count += 1

    # Update filter range
    _format_sheet(ws, SCRAPED_COLUMNS, SCRAPED_HEADER_FILL)
    wb.save(output_path)
    print(f"[Excel] Saved {count} scraped items to '{sheet_name}' sheet")
    return count


def write_events(events: list[dict], output_path: str = DEFAULT_OUTPUT, dry_run: bool = False) -> int:
    """
    Write classified events to the 'Classified Events' sheet with deduplication.
    Returns number of new events added.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    if os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Get or create events sheet
    sheet_name = "Classified Events"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name, 0)  # Insert as first sheet
        _format_sheet(ws, EVENT_COLUMNS, HEADER_FILL)
        _set_column_widths(ws, {
            "Event Name": 35, "Host Company": 18, "Category": 15,
            "Event Type": 14, "Date": 14, "Location": 22,
            "Target Audience": 25, "Registration URL": 40,
            "Description": 50, "Relevance Score": 10, "Source URL": 40,
            "Detected At": 20, "Status": 12,
        }, EVENT_COLUMNS)

    new_count = 0
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    for event in events:
        if dry_run:
            print(f"  [DryRun] Would add: {event.get('event_name')} by {event.get('host_company')}")
            new_count += 1
            continue

        row_idx = ws.max_row + 1
        row_data = [
            event.get("event_name", ""),
            event.get("host_company", ""),
            event.get("source_category", ""),
            event.get("event_type", ""),
            event.get("date", ""),
            event.get("location", ""),
            event.get("target_audience", ""),
            event.get("registration_url", ""),
            event.get("description", ""),
            event.get("relevance_score", 0),
            event.get("source_url", ""),
            now,
            "New",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(name="Calibri", size=10)

        _color_event_type(ws, row_idx, event.get("event_type", ""), EVENT_COLUMNS)

        # Hyperlink for registration URL
        reg_col = EVENT_COLUMNS.index("Registration URL") + 1
        reg_url = event.get("registration_url", "")
        if reg_url and reg_url.startswith("http"):
            cell = ws.cell(row=row_idx, column=reg_col)
            cell.hyperlink = reg_url
            cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

        new_count += 1

    if not dry_run and new_count > 0:
        _format_sheet(ws, EVENT_COLUMNS, HEADER_FILL)
        wb.save(output_path)
        print(f"\n[Excel] Saved {new_count} new events to '{sheet_name}' sheet")
    elif not dry_run:
        print(f"\n[Excel] No new events to add")

    return new_count


# ── LinkedIn Events sheet ──
LINKEDIN_HEADER_FILL = PatternFill(start_color="6B3FA0", end_color="6B3FA0", fill_type="solid")

def write_linkedin_events(events: list[dict], output_path: str = DEFAULT_OUTPUT, dry_run: bool = False) -> int:
    """
    Write LinkedIn-sourced classified events to the 'LinkedIn Events' sheet.
    Separate from web events for clarity.
    Returns number of new events added.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    if os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    sheet_name = "LinkedIn Events"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        _format_sheet(ws, EVENT_COLUMNS, LINKEDIN_HEADER_FILL)
        _set_column_widths(ws, {
            "Event Name": 35, "Host Company": 18, "Category": 15,
            "Event Type": 14, "Date": 14, "Location": 22,
            "Target Audience": 25, "Registration URL": 40,
            "Description": 50, "Relevance Score": 10, "Source URL": 40,
            "Detected At": 20, "Status": 12,
        }, EVENT_COLUMNS)

    new_count = 0
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    for event in events:
        if dry_run:
            print(f"  [DryRun] Would add LinkedIn event: {event.get('event_name')} by {event.get('host_company')}")
            new_count += 1
            continue

        row_idx = ws.max_row + 1
        row_data = [
            event.get("event_name", ""),
            event.get("host_company", ""),
            event.get("source_category", ""),
            event.get("event_type", ""),
            event.get("date", ""),
            event.get("location", ""),
            event.get("target_audience", ""),
            event.get("registration_url", ""),
            event.get("description", ""),
            event.get("relevance_score", 0),
            event.get("source_url", ""),
            now,
            "New (LinkedIn)",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(name="Calibri", size=10)

        _color_event_type(ws, row_idx, event.get("event_type", ""), EVENT_COLUMNS)

        reg_col = EVENT_COLUMNS.index("Registration URL") + 1
        reg_url = event.get("registration_url", "")
        if reg_url and reg_url.startswith("http"):
            cell = ws.cell(row=row_idx, column=reg_col)
            cell.hyperlink = reg_url
            cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

        new_count += 1

    if not dry_run and new_count > 0:
        _format_sheet(ws, EVENT_COLUMNS, LINKEDIN_HEADER_FILL)
        wb.save(output_path)
        print(f"\n[Excel] Saved {new_count} LinkedIn events to '{sheet_name}' sheet")
    elif not dry_run:
        print(f"\n[Excel] No new LinkedIn events to add")

    return new_count


# ── Engagers sheet ──
ENGAGER_COLUMNS = [
    "Name",
    "LinkedIn URL",
    "Title",
    "Company",
    "Email",
    "Company Size",
    "Industry",
    "Location",
    "ICP Score",
    "Verdict",
    "Engagement Type",
    "Engagement Strength",
    "Event Date",
    "Event City",
    "Warm Reason",
    "Source Post",
    "Source Company",
    "Enriched At",
]

ENGAGER_HEADER_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")


def write_engagers(engagers: list[dict], output_path: str = DEFAULT_OUTPUT, dry_run: bool = False) -> int:
    """
    Write enriched engagers to the 'Engagers' sheet.
    Returns number of engagers written.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    if os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    sheet_name = "Engagers"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        _format_sheet(ws, ENGAGER_COLUMNS, ENGAGER_HEADER_FILL)
        _set_column_widths(ws, {
            "Name": 22, "LinkedIn URL": 40, "Title": 25,
            "Company": 22, "Email": 30, "Company Size": 14,
            "Industry": 20, "Location": 22, "ICP Score": 10,
            "Verdict": 16, "Engagement Type": 14, "Engagement Strength": 16,
            "Event Date": 12, "Event City": 10, "Warm Reason": 50,
            "Source Post": 40, "Source Company": 18, "Enriched At": 20,
        }, ENGAGER_COLUMNS)

    new_count = 0
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    for engager in engagers:
        if dry_run:
            score = engager.get("icp_score", "?")
            print(f"  [DryRun] Would add engager: {engager.get('name')} "
                  f"({engager.get('parsed_title', '?')}) [ICP: {score}]")
            new_count += 1
            continue

        row_idx = ws.max_row + 1
        linkedin_url = engager.get("linkedin_url", "")
        row_data = [
            engager.get("name", ""),
            linkedin_url,
            engager.get("title", "") or engager.get("parsed_title", ""),
            engager.get("company", "") or engager.get("parsed_company", ""),
            engager.get("email", ""),
            engager.get("company_size", ""),
            engager.get("industry", ""),
            engager.get("location", ""),
            engager.get("icp_score", 0),
            engager.get("verdict", ""),
            engager.get("engagement_type", ""),
            engager.get("engagement_strength", ""),
            engager.get("event_date", ""),
            engager.get("event_city", ""),
            engager.get("warm_reason", ""),
            engager.get("source_post_url", ""),
            engager.get("source_post_company", ""),
            now,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(name="Calibri", size=10)

        # Hyperlink for LinkedIn URL
        if linkedin_url and linkedin_url.startswith("http"):
            cell = ws.cell(row=row_idx, column=2)
            cell.hyperlink = linkedin_url
            cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

        # Color code ICP score
        icp_score = engager.get("icp_score", 0)
        score_cell = ws.cell(row=row_idx, column=ENGAGER_COLUMNS.index("ICP Score") + 1)
        if icp_score >= 70:
            score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            score_cell.font = Font(name="Calibri", size=10, color="1E6B22", bold=True)
        elif icp_score >= 50:
            score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            score_cell.font = Font(name="Calibri", size=10, color="7D5700", bold=True)
        else:
            score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            score_cell.font = Font(name="Calibri", size=10, color="9C0006", bold=True)

        new_count += 1

    if not dry_run and new_count > 0:
        _format_sheet(ws, ENGAGER_COLUMNS, ENGAGER_HEADER_FILL)
        wb.save(output_path)
        print(f"\n[Excel] Saved {new_count} engagers to '{sheet_name}' sheet")
    elif not dry_run:
        print(f"\n[Excel] No engagers to add")

    return new_count


if __name__ == "__main__":
    test_events = [
        {
            "event_name": "CFO Summit 2026",
            "host_company": "Ramp",
            "source_category": "accounts_payable",
            "event_type": "Summit",
            "date": "2026-03-15",
            "location": "San Francisco, CA",
            "target_audience": "CFOs, Controllers, VP Finance",
            "registration_url": "https://example.com/cfo-summit",
            "description": "A networking dinner for CFOs at B2B SaaS companies.",
            "relevance_score": 9,
            "source_url": "https://ramp.com/events",
        },
    ]
    count = write_events(test_events)
    print(f"Added {count} events")

